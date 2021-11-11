from openpyxl import load_workbook

def dbxl_connect():
    wb = load_workbook('db.xlsx')
    sheet = wb.active
    return sheet
    
def get_value(cordiate):
    sheet = dbxl_connect()
    value = sheet[cordiate].value
    return value
    
def get_num_user(message):
    if type(message) == int:
        id = str(message)   # Dars qo'shishda id son ko'rinishida keladi va shundan ajratib oladi        
    else:
        id = str(message.chat.id)
    
    sheet = dbxl_connect()
    IDs = []
    
    for row in sheet.rows:
        IDs.append(str(row[0].value))
    
    if id not in IDs:
        N = len(IDs)
    else:
        N = IDs.index(id) + 1
    
    return N
    
def add_user(message):
    id = str(message.chat.id)
    fullname = message.from_user.full_name

    wb = load_workbook('db.xlsx') 
    sheet = wb.active
    
    IDs = []
    for row in sheet.rows:
        IDs.append(str(row[0].value))
    
    if id not in IDs:
        N = len(IDs) + 1
        sheet[f'A{N}'] = id
        sheet[f'B{N}'] = fullname
        
    wb.save('db.xlsx')

def add_data(cordinate, item):
    wb = load_workbook('db.xlsx')
    sheet = wb.active
    sheet[cordinate] = item
    wb.save('db.xlsx')

def get_dars(message, day):
    sheet = dbxl_connect()
    N = get_num_user(message)
    darslar = sheet[f"{day}{N}"].value
    if darslar != None:
        return darslar
    else:
        kun = sheet[f"{day}1"].value
        kun = kun.title()
        
        text = f"{kun} kunlik dars jadvali kiritilmagan"
        return text

def get_data(cordinate):
    sheet = dbxl_connect()
    data = sheet[cordinate].value
    return data

def get_cordinate(N, item):
    let_list = ['D', 'E', 'F', 'J', 'K', 'L', 'M', 'N']
    sheet = dbxl_connect()
    
    for let in let_list:
        if sheet[f"{let}{N}"].value == item:
            return f"{let}{N}"